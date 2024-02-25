from datetime import datetime
import os
from msal import ConfidentialClientApplication
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from inApp.models import Orden, DetallesOrden
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.creation_information import FileCreationInformation
from datetime import date

def get_sharepoint_context_using_app():
    # Get SharePoint credentials
    sharepoint_url = 'https://distribuidoraagroshopspa.sharepoint.com'

    # Initialize the client credentials
    client_credentials = ClientCredential('be6f1624-be6b-4ce3-85eb-dc7302278df3', 'nXx8Q~gJQa0BcTjZHKkkwCVxTtCSSH1yNXmN2bE~')

    # Create client context object
    ctx = ClientContext(sharepoint_url).with_credentials(client_credentials)

    return ctx

def upload_file_to_sharepoint(ctx, folder_url, file_name, file_content):
    # Create a FileCreationInformation object
    file_info = FileCreationInformation()
    file_info.url = file_name
    file_info.content = file_content

    # Get the folder to upload the file to
    folder = ctx.web.get_folder_by_server_relative_url(folder_url)
    ctx.load(folder)
    ctx.execute_query()

    # Upload the file to the folder
    uploaded_file = folder.files.add(file_info.url, file_info.content, overwrite=True)
    ctx.execute_query()

    print(f"File '{file_name}' uploaded successfully to SharePoint.")

def agregar_datos_a_excel(request, orden):
    print("Iniciando la función agregar_datos_a_excel")

    # Configuración de SharePoint
    ctx = get_sharepoint_context_using_app()

    # Ruta completa al archivo pedidos.xlsx en SharePoint
    today = datetime.today().strftime('%d-%m')
    directory_path = "Documentos compartidos/archivos_excel/"
    path = os.path.join(directory_path, f"pedidos_{today}.xlsx")

    # Cargar el libro de trabajo si ya existe, de lo contrario, crear uno nuevo
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()

    # Obtén la hoja correspondiente al día actual o crea una nueva
    sheet_name = today
    if sheet_name not in wb.sheetnames:
        hoja = wb.create_sheet(sheet_name)
        
        # Establecer el tamaño de la fuente para todas las celdas
        for row in hoja.iter_rows():
            for cell in row:
                cell.font = Font(size=18)

        # Encabezados (agregar solo si la hoja es nueva)
        if not hoja['A1'].value or not hoja['B1'].value or not hoja['C1'].value:
            hoja['A1'] = 'Producto'
            hoja['B1'] = 'SKU'
            hoja['C1'] = 'Unidad'
    else:
        # Si la hoja ya existe, seleccionarla
        hoja = wb[sheet_name]





    usuario_actual = orden.usuario.username

    if usuario_actual not in [hoja.cell(row=1, column=col).value for col in range(2, hoja.max_column + 1)]:
        # Agregar el nombre de usuario como columna
        col_usuario = hoja.max_column + 1
        hoja.cell(row=1, column=col_usuario).value = usuario_actual


    letra_columna_productos = get_column_letter(1)
    letra_columna_sku = get_column_letter(2)
    letra_columna_unidad = get_column_letter(3)
    

    # Ajustar el tamaño de la celda para la columna de productos
    hoja.column_dimensions[letra_columna_productos].width = 60
    for row in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(size=18)  # Adjust font size for the entire column

    # Ajustar el tamaño de la celda para la columna de unidad
    hoja.column_dimensions[letra_columna_sku].width = 20
    for row in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.font = Font(size=18)  # Adjust font size for the entire column

    hoja.column_dimensions[letra_columna_unidad].width = 18
    for row in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.font = Font(size=18)  # Adjust font size for the entire column




        








 




    ordenes_en_hoja = [hoja.cell(row=fila, column=1).value for fila in range(2, hoja.max_row + 1) if hoja.cell(row=fila, column=1).value]


    col_usuario = [col for col in range(3, hoja.max_column + 1) if hoja.cell(row=1, column=col).value == usuario_actual]

    if not col_usuario:
        # Si el usuario no tiene columna, agregar una nueva
        col_usuario = hoja.max_column + 1
        hoja.cell(row=1, column=col_usuario).value = usuario_actual
    else:
        col_usuario = col_usuario[0]
        

    for col in [col_usuario]:
        letra_columna = get_column_letter(col)
        for cell in hoja[letra_columna]:
            cell.font = Font(size=18)

            

# ...

# ...

# ...




    fecha_actual = date.today()
    
             

    ordenes_confirmadas = Orden.objects.filter(compra_confirmada=True, fecha=fecha_actual)

    for orden_actual in ordenes_confirmadas:
        detalles = DetallesOrden.objects.filter(orden=orden_actual)
        for detalle in detalles:

            for detalle in detalles:
                producto = detalle.producto.nombre
                sku = detalle.producto.sku 
                cantidad = detalle.cantidad
                unidad = detalle.unidad
                print(f"Producto: {producto}, SKU: {sku}, Cantidad: {cantidad}, Unidad: {unidad}")

                # Encontrar la primera fila vacía para agregar los datos
                fila_producto = next((fila for fila in range(2, hoja.max_row + 1) if hoja.cell(row=fila, column=1).value == producto), None)

                if fila_producto is None:
                    # Si el producto no existe, encontrar la primera fila vacía para agregar los datos
                    fila_producto = 2
                    while hoja.cell(row=fila_producto, column=2).value:
                        fila_producto += 1
                    # Agregar datos a la hoja
                    hoja.cell(row=fila_producto, column=1).value = producto

                # Si el producto ya existe, no es necesario agregar una nueva fila
                # Puedes decidir qué hacer en este caso, por ejemplo, sumar las cantidades o ignorar la orden
                # Aquí estoy sumando las cantidades para ejemplificar
                cantidad_existente = hoja.cell(row=fila_producto, column=col_usuario).value or 0
                cantidad_existente += cantidad
                hoja.cell(row=fila_producto, column=col_usuario).value = cantidad_existente

                hoja.cell(row=fila_producto, column=2).value = sku  # Assuming SKU column is at index 3



                # Agregar unidad a la hoja debajo de "Unidad"
                hoja.cell(row=fila_producto, column=3).value = unidad

                

                # Agregar cantidad a la hoja debajo del usuario
                hoja.cell(row=fila_producto, column=col_usuario).value = cantidad

                






    col_hola = next((col for col in range(2, hoja.max_column + 1) if hoja.cell(row=1, column=col).value == 'totale'), None)

    if col_hola is not None:
        # Eliminar la columna "hola" si ya existe
        hoja.delete_cols(col_hola)

    # Añadir la columna "hola" al final
    col_hola = hoja.max_column + 1
    hoja.cell(row=1, column=col_hola).value = 'totale'

    hoja.cell(row=1, column=col_hola).fill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')

    hoja.cell(row=1, column=col_hola).font = Font(size=18)
    # Ajustar el tamaño de la celda
    letra_columna_hola = get_column_letter(col_hola)
    hoja.column_dimensions[letra_columna_hola].width = 20

    hoja.auto_filter.ref = hoja.dimensions
    hoja.auto_filter.add_filter_column(col_hola, ["totale"])
    



    for fila in range(2, hoja.max_row + 1):
      if hoja.cell(row=fila, column=1).value:
        total_producto = sum(
            float(hoja.cell(row=fila, column=col).value or 0)
            for col in range(4, hoja.max_column + 1)  # Ajustar el rango hasta la última columna
        )
        hoja.cell(row=fila, column=col_hola).value = total_producto


    col_inventario = next((col for col in range(2, hoja.max_column + 1) if hoja.cell(row=1, column=col).value == 'inventario'), None)

    if col_inventario is not None:
        # Eliminar la columna "inventario" si ya existe
        hoja.delete_cols(col_inventario)

    # Añadir la columna "inventario" al final
    col_inventario = hoja.max_column + 1
    hoja.cell(row=1, column=col_inventario).value = 'inventario'

    hoja.cell(row=1, column=col_inventario).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    hoja.cell(row=1, column=col_inventario).font = Font(size=18)
    # Ajustar el tamaño de la celda
    letra_columna_inventario = get_column_letter(col_inventario)
    hoja.column_dimensions[letra_columna_inventario].width = 20

    hoja.auto_filter.ref = hoja.dimensions
    hoja.auto_filter.add_filter_column(col_inventario, ["inventario"])



   


    col_por_comprar = next((col for col in range(2, hoja.max_column + 1) if hoja.cell(row=1, column=col).value == 'por comprar'), None)

    if col_por_comprar is not None:
        # Eliminar la columna "por comprar" si ya existe
        hoja.delete_cols(col_por_comprar)

    # Añadir la columna "por comprar" al final
    col_por_comprar = hoja.max_column + 1
    hoja.cell(row=1, column=col_por_comprar).value = 'por comprar'

    hoja.cell(row=1, column=col_por_comprar).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    hoja.cell(row=1, column=col_por_comprar).font = Font(size=18)
    # Ajustar el tamaño de la celda
    letra_columna_por_comprar = get_column_letter(col_por_comprar)
    hoja.column_dimensions[letra_columna_por_comprar].width = 20

    hoja.auto_filter.ref = hoja.dimensions
    hoja.auto_filter.add_filter_column(col_por_comprar, ["por comprar"])




    col_proveedor = next((col for col in range(2, hoja.max_column + 1) if hoja.cell(row=1, column=col).value == 'proveedor'), None)

    if col_proveedor is not None:
        # Eliminar la columna "proveedor" si ya existe
        hoja.delete_cols(col_proveedor)

    # Añadir la columna "proveedor" al final
    col_proveedor = hoja.max_column + 1
    hoja.cell(row=1, column=col_proveedor).value = 'proveedor'

    hoja.cell(row=1, column=col_proveedor).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    hoja.cell(row=1, column=col_proveedor).font = Font(size=18)
    # Ajustar el tamaño de la celda
    letra_proveedor = get_column_letter(col_proveedor)
    hoja.column_dimensions[letra_proveedor].width = 20

    hoja.auto_filter.ref = hoja.dimensions
    hoja.auto_filter.add_filter_column(col_proveedor, ["proveedor"])



    for fila in range(1, hoja.max_row + 1):
        hoja.cell(row=fila, column=1).fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

                # Aplicar formato de relleno al color especificado (RGB: 238, 236, 225) a toda la columna de unidad
    for fila in range(1, hoja.max_row + 1):
        hoja.cell(row=fila, column=2).fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

    for fila in range(1, hoja.max_row + 1):
        hoja.cell(row=fila, column=3).fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

    borde = Border(
        left=Side(border_style='thin', color='000000'),  # Borde izquierdo
        right=Side(border_style='thin', color='000000'),  # Borde derecho
        top=Side(border_style='thin', color='000000'),    # Borde superior
        bottom=Side(border_style='thin', color='000000')  # Borde inferior
    )

    # Aplicar bordes a todas las celdas de la hoja
    for row in hoja.iter_rows():
        for cell in row:
            cell.border = borde
                


    try:
        wb.save(path)
        print(f"Archivo guardado correctamente localmente en la ruta: {path}")

        # Cargar el archivo en SharePoint
        with open(path, 'rb') as file_content:
            folder_url = "/Documentos compartidos/archivos_excel"  # Actualiza con la ruta correcta en tu SharePoint
            file_name = f"pedidos_{today}.xlsx"

            # Llamar a la función para cargar el archivo en SharePoint
            upload_file_to_sharepoint(ctx, folder_url, file_name, file_content)

        print(f"Archivo cargado correctamente en SharePoint en la carpeta: {folder_url}")

    except Exception as e:
        print(f"Error al guardar o cargar el archivo: {e}")
